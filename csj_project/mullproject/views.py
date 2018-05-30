# coding=utf-8

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from mullproject.models import Settlement, Category, Question
from mullproject.forms import SettlementForm, CategoryForm, QuestionForm
from mullproject.serializers import SettlementSerializer, QuestionSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.renderers import JSONRenderer
from openpyxl import load_workbook
from zipfile import BadZipfile
from io import BytesIO
import os, csv, re

# APIs for Unity map
class SettlementList(APIView):
    renderer_classes = (JSONRenderer, )
    def get(self, request, format=None):
         settlements = Settlement.objects.all()
         serializer = SettlementSerializer(settlements, many=True)
         return Response(serializer.data)

class QuestionList(APIView):
    renderer_classes = (JSONRenderer, )
    def get(self, request, format=None):
         questions = Question.objects.all()
         serializer = QuestionSerializer(questions, many=True)
         return Response(serializer.data)

# index/home page
def index(request):
    return render(request, 'mullproject/index.html')

# map with game
def game(request):
    return render(request, 'mullproject/game.html')

# map without game
def map(request):
    return render(request, 'mullproject/map.html')

# list of all categories
def categories(request):
    category_list = Category.objects.order_by('gaelic_name')
    context_dict = {'categories':category_list}
    return render(request, 'mullproject/categories.html', context_dict)

# list of all settlements contained within given 'category'
def settlements(request, category):
    try:
        cat = Category.objects.get(slug=category)
    except Category.DoesNotExist:
        cat = None
    settlement_list = Settlement.objects.filter(categories=category).order_by('headname')

    settlements = None

    if cat:
        # split settlement into pages of five
        num_per_page = 5
        paginator = Paginator(settlement_list, num_per_page)
        page = request.GET.get('page')
        try:
            settlements = paginator.page(page)
        except PageNotAnInteger:
            settlements = paginator.page(1)
        except EmptyPage:
            settlements = paginator.page(paginator.num_pages)

    context_dict = {'settlements':settlements, 'category':cat}
    return render(request, 'mullproject/settlements.html', context_dict)

# an individual settlement and it's information
def settlement(request, settlement):
    try:
        settlement = Settlement.objects.get(slug=settlement)
    except Settlement.DoesNotExist:
        settlement = None
    categories = None
    if settlement:
        categories = Category.objects.filter(settlement=settlement)
    context_dict = {'settlement': settlement, 'categories': categories}
    return render(request, 'mullproject/settlement.html', context_dict)

# list of all categories and settlements and option to delete them
def manage(request):
    categories = Category.objects.order_by('gaelic_name')
    settlements = Settlement.objects.order_by('headname')
    questions = Question.objects.all()

    unsuitableFileType = False
    fileUploaded = False

    #Checks if a file has been uploaded and adds to database
    if request.POST and request.FILES:

        try:
	       #Get hold of the file that has been uploaded
            file_in_memory = request.FILES['csv_file'].read()
            wb = load_workbook(filename=BytesIO(file_in_memory))
            sheet=wb['Sheet1']

	        #Ignore the first row of the file (headers)
            i = 2

	        #Process each row of the file until it hits an empty row
            while(sheet.cell(row=i,column=1).value != None):

	            #Get the english version of language of origin
                langOfOrigin = sheet.cell(row=i,column=5).value.split("|")[1]

	            #Create a list of categories with Gaelic and English versions
                uploadedCategories = sheet.cell(row=i,column=10).value
                uploadedCategories = uploadedCategories.replace(';','|').split("|")

	            #Create categories for Gaelic, Old Norse and Gaelic/Old Norse
                Category.objects.get_or_create(gaelic_name = "Gàidhlig", english_name = "Gaelic")
                Category.objects.get_or_create(gaelic_name = "Seann Lochlannais", english_name = "Old Norse")
                englishCategories = []

	            #Clean up language of origin so we are left with it in the same form as model
                langOfOrigin = langOfOrigin.strip().split()
                if(len(langOfOrigin) > 1 and langOfOrigin[1] == "(?)"):
                    langOfOrigin[1] = ""
                langOfOrigin = " ".join(langOfOrigin).strip()

	            #Add the correct Language of Origin to the categories list for the current settlement
                if(langOfOrigin.split()[0].replace(' ','') == "Gaelic"):
                    englishCategories.append("Gaelic")
                elif(langOfOrigin.split()[0].replace(' ','') == "Old" and langOfOrigin.split()[1].replace(' ','') == "Norse"):
                    englishCategories.append("Old Norse")
                elif(langOfOrigin.split()[0].replace(' ','') == "Gaelic/Old" and langOfOrigin.split()[1].replace(' ','') == "Norse"):
                    englishCategories.append("Gaelic")
                    englishCategories.append("Old Norse")

	            #Add each categories from the uploaded file to the list of categories
	            #Create a Category object for each of these
                for j in range(0, (len(uploadedCategories)/2)):
                    englishCategories.append(uploadedCategories[j+len(uploadedCategories)/2])
                    try:
                        gaelicCatName = uploadedCategories[j].lstrip()
                        englishCatName = uploadedCategories[j+len(uploadedCategories)/2]
                        Category.objects.get_or_create(gaelic_name = gaelicCatName, english_name = englishCatName)

                    except Exception, e:
		                print(str(e))

	            #Create a settlement object for each of the rows in the file
                try:
			        s = Settlement.objects.create(headname = sheet.cell(row=i,column=1).value,
				                                  anglicised=sheet.cell(row=i,column=3).value,
			    	                              grid_ref = sheet.cell(row=i,column=2).value,
				                                  historical_forms = sheet.cell(row=i,column=4).value,
				                                  original_elements = sheet.cell(row=i,column=6).value,
			    	                              interpretation = sheet.cell(row=i,column=7).value,
				                                  extent = sheet.cell(row=i,column=8).value,
				                                  certainty = sheet.cell(row=i,column=9).value,
			    	                              lang_of_origin = sheet.cell(row=i,column=5).value.split("|")[1])
                except Exception, e:
				    print(str(e))

	            #Link the categories and settlements together
                try:
                    for cat in englishCategories:
			            s.categories.add(Category.objects.get(english_name = cat))
                except Exception, e:
				    print(str(e))



	            #Create default questions for each settlement
                try:
                    headname = sheet.cell(row=i,column=1).value
                    englishText = "Find the settlement with headname " + headname
                    gaelicText = "Lorg an tuineachadh le ceann-ainm " + headname
                    Question.objects.get_or_create(text_english=englishText,
                                                   text_gaelic=gaelicText,
                                                   answer=Settlement.objects.get(headname=headname)
                                                   )
                    interpretation = sheet.cell(row=i,column=7).value
                    englishText = "Find the settlement which translates to " + interpretation
                    gaelicText = "Lorg an tuineachadh a tha ag eadar-theangachadh " + interpretation
                    Question.objects.get_or_create(text_english=englishText,
                                                   text_gaelic=gaelicText,
                                                   answer=Settlement.objects.get(headname=headname)
                                                   )
                except Exception, e:
				    print(str(e))

                i += 1

            fileUploaded = True

        except BadZipfile:
            unsuitableFileType = True
        except IOError:
            unsuitableFileType = True

    context_dict = {'categories': categories,
                    'settlements': settlements,
                    'questions': questions,
                    'unsuitableFileType': unsuitableFileType,
                    'fileUploaded': fileUploaded,
                    }
    return render(request, 'mullproject/manage.html', context_dict)

# form to add a new category
def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)

        # if valid, save and redirect
        if form.is_valid():
            form.save(commit=True)
            return redirect('manage')

        # if invalid, tell user why
        else:
            print form.errors
    else:
        form = CategoryForm()
    context_dict = {'form': form}
    return render(request, 'mullproject/add_category.html', context_dict)

# form to add a new settlement
def add_settlement(request):
    if request.method == 'POST':
        form = SettlementForm(request.POST)

        # if valid, save and redirect
        if form.is_valid():
            form.save(commit=True)
            return redirect('manage')

        # if invalid, tell user why
        else:
            print form.errors
    else:
        form = SettlementForm()
        context_dict = {'form': form}
        return render(request, 'mullproject/add_settlement.html', context_dict)

# form to add new category
def add_question(request):
    if request.method == 'POST':
        form = QuestionForm(request.POST)

        #if valid, save and redirect
        if form.is_valid():
            form.save(commit=True)
            return redirect('manage')

        #if invalid, tell user why
        else:
            print form.errors

    else:
        form = QuestionForm()
        context_dict = {'form': form}
        return render(request, 'mullproject/add_question.html', context_dict)

# confirms that a category has been deleted
def delete_category(request, category):
    deleted = False

    # get the category to be deleted
    try:
        cat = Category.objects.get(slug=category)
    except Category.DoesNotExist:
        cat = None

    # get any settlements in this category
    # if there is only one, settlements is that settlement
    try:
        settlements = Settlement.objects.filter(categories=category).order_by('headname')

    # if there are none, settlements is false
    except Settlement.DoesNotExist:
        settlements = None

    # if the category exists and there are no settlements in it, delete it
    if cat and not settlements:
        cat.delete()
        deleted = True
    context_dict = {'cat': cat, 'deleted': deleted, 'settlements': settlements}
    return render(request, 'mullproject/delete_category.html', context_dict)

# confirms that a settlement has been deleted
def delete_settlement(request, settlement):

    # get the settlement to be deleted
    try:
        s = Settlement.objects.get(slug=settlement)
    except Settlement.DoesNotExist:
        s = None

    # if it exists, delete it
    if s:
        os.remove('csj_project/static/media/settlement_images/' + s.slug + '.png')
        s.delete()
    context_dict = {'settlement': s}
    return render(request, 'mullproject/delete_settlement.html', context_dict)

# confirms that a question has been deleted
def delete_question(request, question):

    #get the question to be deleted
    try:
        q = Question.objects.get(id=question)
    except Question.DoesNotExist:
        q = None
    except ValueError:
        q = None

    #if it exists, delete it
    if q:
        q.delete()
    context_dict = {'question': q}
    return render(request, 'mullproject/delete_question.html', context_dict)

# allows a user to login
def user_login(request):
    if request.user.is_authenticated():
        return redirect('manage')
    else:
        failed = False
        if request.method == 'POST':

            # get the input username and password
            username = request.POST.get('username')
            password = request.POST.get('password')

            # check that these details are correct
            user = authenticate(username=username, password=password)

            # if correct, log them in and take them to 'manage.html'
            if user:
                login(request, user)
                return redirect('manage')

            # if incorrect, display error message
            else:
                failed = True
                context_dict = {'failed':failed}
                return render(request, 'mullproject/login.html', context_dict)

        else:
            return render(request, 'mullproject/login.html')

# logs a user out and redirects to index
def user_logout(request):
    logout(request)
    return redirect('index')

# about us page
def about(request):
    return render(request, 'mullproject/about.html')
